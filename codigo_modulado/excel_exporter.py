import openpyxl
from openpyxl.styles import Font, Alignment

def exportar_seguidores_excel(nombre_archivo, usuario_objetivo, seguidores_info):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Seguidores de @{usuario_objetivo}"
    # Título
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Seguidores de @{usuario_objetivo}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    # Encabezados
    ws.append(["Seguidor", "Fecha de unión", "Link del perfil"])
    ws['A2'].font = ws['B2'].font = ws['C2'].font = Font(bold=True)
    # Datos
    for username, fecha in seguidores_info:
        ws.append([
            username,
            fecha if fecha else "No encontrada",
            f"https://www.instagram.com/{username}/"
        ])
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    wb.save(nombre_archivo)
    print(f"\n📁 Archivo Excel guardado como: {nombre_archivo}")
