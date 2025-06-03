from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from dotenv import load_dotenv
import os
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
import random
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
# Cargar variables de entorno
load_dotenv()

USERNAME = "nico.proyecx6@gmail.com"
PASSWORD = os.getenv('PASSWORD')

# Solicitar el usuario a buscar
print("🔍 INSTAGRAM FOLLOWERS SCRAPER")
print("="*50)
usuario_objetivo = input("👤 Ingresa el nombre de usuario de Instagram a buscar (sin @): ").strip()

# Validar entrada
if not usuario_objetivo:
    print("❌ Error: Debes ingresar un nombre de usuario válido")
    exit()

# Remover @ si lo incluyó el usuario
if usuario_objetivo.startswith('@'):
    usuario_objetivo = usuario_objetivo[1:]

print(f"🎯 Buscando seguidores de: @{usuario_objetivo}")
print("="*50)

# Ruta correcta del chromedriver
chromedriver_path = "C:\\Users\\User\\OneDrive - Universidad Nacional de Colombia\\Escritorio\\codigo_Prueba\\chromedriver\\chromedriver-win64\\chromedriver.exe"

def obtener_fecha_union(driver, username):
    """
    Obtiene la fecha de unión de un usuario a Instagram
    """
    try:
        perfil_url = f"https://www.instagram.com/{username}/"
        print(f"📅 Obteniendo fecha de unión de @{username}...")
        driver.get(perfil_url)
        sleep(random.uniform(2, 3))

        # Buscar el botón de opciones
        boton_opciones = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Options' or @aria-label='Opciones']"))
        )
        boton_opciones.click()
        sleep(random.uniform(1.5, 2.5))

        # Buscar la fecha de unión
        fecha_xpath = "//span[contains(text(),'Fecha en la que te uniste')]/following-sibling::span"
        fecha_element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, fecha_xpath))
        )
        fecha = fecha_element.text
        
        # Cerrar el menú de opciones haciendo clic fuera
        driver.execute_script("document.body.click();")
        sleep(1)
        
        return fecha
    except Exception as e:
        print(f"⚠️ No se pudo obtener fecha de @{username}: {str(e)[:50]}...")
        
        # Intentar cerrar cualquier modal abierto
        try:
            driver.execute_script("document.body.click();")
            sleep(1)
        except:
            pass
            
        return "Fecha no encontrada"

def es_username_valido(username):
    """
    Valida si un username es válido según las reglas de Instagram
    y no es parte de las páginas especiales de Instagram
    """
    if not username or len(username) == 0:
        return False
    
    # Lista completa de páginas especiales y patrones a excluir
    paginas_especiales = {
        # Páginas principales de Instagram
        'explore', 'reels', 'tv', 'stories', 'accounts', 'direct', 'about',
        'followers', 'following', 'inbox', 'locations', 'lite',
        'help', 'legal', 'privacy', 'safety', 'api', 'blog',
        'business', 'community', 'developers', 'support', 'press',
        'careers', 'brand', 'download', 'create', 'features',
        
        # Parámetros especiales
        '?entrypoint=web_footer', 'entrypoint=web_footer'
    }
    
    # Normalizar username (remover caracteres especiales para la comparación)
    username_clean = username.lower().strip()
    
    # 1. Verificar si es una página especial
    if username_clean in paginas_especiales:
        return False
    
    # 2. Verificar si contiene parámetros de URL
    if '?' in username or '&' in username or '=' in username:
        return False
    
    # 3. Verificar patrones de IDs temporales de Instagram (como DJ4EA7ZIaj7)
    # Estos suelen tener patrones específicos: 2-3 letras seguidas de caracteres alfanuméricos
    patron_id_temporal = re.compile(r'^[A-Z]{2,3}[A-Za-z0-9_-]{6,}$')
    if patron_id_temporal.match(username):
        return False
    
    # 4. Verificar usernames que son solo números o caracteres extraños
    if username.isdigit():
        return False
    
    # 5. Verificar longitud mínima y máxima (Instagram permite 1-30 caracteres)
    if len(username) < 1 or len(username) > 30:
        return False
    
    # 6. Verificar que no contenga solo caracteres especiales
    if re.match(r'^[._-]+$', username):
        return False
    
    # 7. Verificar que no empiece o termine con punto
    if username.startswith('.') or username.endswith('.'):
        return False
    
    # 8. Verificar que no contenga espacios
    if ' ' in username:
        return False
    
    # 9. Verificar que solo contenga caracteres válidos para Instagram
    # Instagram permite: letras, números, puntos, guiones bajos
    patron_valido = re.compile(r'^[a-zA-Z0-9._]+$')
    if not patron_valido.match(username):
        return False
    
    # 10. Verificar que no sea el propio usuario objetivo
    if username.lower() == usuario_objetivo.lower():
        return False
    
    return True

def limpiar_username(link):
    """
    Extrae y limpia el username de un enlace de Instagram
    """
    try:
        if not link or not isinstance(link, str):
            return None
        
        # Remover protocolo y dominio si están presentes
        if 'instagram.com/' in link:
            username = link.split('instagram.com/')[-1]
        else:
            username = link
        
        # Remover barra final y parámetros
        username = username.rstrip('/').split('?')[0].split('&')[0]
        
        # Remover barras adicionales (en caso de que sea parte de una ruta)
        if '/' in username:
            username = username.split('/')[0]
        
        return username.strip()
    except:
        return None

# Función para inicializar el driver
def iniciar_driver():
    options = Options()
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/136.0.0.0 Safari/537.36")
    # Agregar argumentos adicionales para mejor estabilidad
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    service = Service(chromedriver_path)

    try:
        driver = webdriver.Chrome(service=service, options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        return driver
    except Exception as e:
        print(f"Error al iniciar el driver: {e}")
        exit()

def navegar_a_perfil_desde_inicio(driver, usuario_objetivo):
    """
    Navega al perfil del usuario desde la página principal de Instagram
    """
    try:
        print(f"🏠 Volviendo al inicio de Instagram...")
        driver.get('https://www.instagram.com/')
        sleep(random.uniform(3, 5))
        
        print(f"📱 Navegando al perfil de @{usuario_objetivo}...")
        profile_url = f'https://www.instagram.com/{usuario_objetivo}/'
        driver.get(profile_url)
        sleep(random.uniform(4, 6))
        
        # Verificar que el perfil cargó correctamente
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "header"))
        )
        print(f"✅ Perfil de @{usuario_objetivo} cargado correctamente")
        return True
        
    except Exception as e:
        print(f"❌ Error al navegar al perfil: {e}")
        return False

def encontrar_boton_seguidores(driver, usuario_objetivo, max_intentos=3):
    """
    Intenta encontrar el botón de seguidores usando diferentes métodos
    Si falla, vuelve al inicio y reintenta
    """
    estrategias = [
        # Estrategia 1: Por href que contenga followers (más confiable)
        (By.XPATH, "//a[contains(@href, '/followers/')]"),
        
        # Estrategia 2: Por texto que contenga números + "followers" o "seguidores"
        (By.XPATH, "//a[contains(text(), 'followers') or contains(text(), 'seguidores')]"),
        
        # Estrategia 3: Estructura típica de Instagram - segundo li en stats
        (By.XPATH, "//header//ul/li[2]//a"),
        
        # Estrategia 4: Por posición - segundo elemento clickeable en stats
        (By.XPATH, "(//header//ul/li//a)[2]"),
        
        # Estrategia 5: CSS Selector equivalente
        (By.CSS_SELECTOR, "header ul li:nth-child(2) a"),
        
        # Estrategia 6: Más específico para la estructura actual
        (By.XPATH, "//section/main/div/header//ul/li[2]/div/a"),
    ]
    
    for intento in range(max_intentos):
        print(f"\n🔄 Intento {intento + 1}/{max_intentos} para encontrar botón de seguidores")
        
        # Si no es el primer intento, volver al inicio y navegar de nuevo
        if intento > 0:
            print("🔄 Volviendo al inicio de Instagram antes del nuevo intento...")
            if not navegar_a_perfil_desde_inicio(driver, usuario_objetivo):
                continue
        
        # Intentar cada estrategia
        for i, (by, selector) in enumerate(estrategias, 1):
            try:
                print(f"Intentando estrategia {i}: {selector}")
                element = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((by, selector))
                )
                # Verificar que el elemento sea realmente de seguidores
                href = element.get_attribute('href')
                text = element.text
                if href and 'followers' in href:
                    print(f"✅ Estrategia {i} exitosa. Elemento encontrado: {text}")
                    return element
                elif 'followers' in text.lower() or 'seguidores' in text.lower():
                    print(f"✅ Estrategia {i} exitosa por texto. Elemento encontrado: {text}")
                    return element
            except Exception as e:
                print(f"❌ Estrategia {i} falló: {str(e)[:100]}")
                continue
        
        # Si llegamos aquí, todas las estrategias fallaron en este intento
        print(f"❌ Intento {intento + 1} fallido. Todas las estrategias fallaron.")
        if intento < max_intentos - 1:
            print(f"⏳ Esperando antes del siguiente intento...")
            sleep(random.uniform(5, 8))
    
    return None

def cargar_seguidores_completo(driver, usuario, limite_max=2000, max_intentos_carga=2):
    """
    Carga seguidores haciendo scroll progresivo hasta alcanzar el límite o no encontrar más
    Si falla durante la carga, vuelve al inicio y reintenta
    VERSIÓN MEJORADA CON OBTENCIÓN DE FECHAS DE UNIÓN
    """
    print(f"📜 Iniciando carga de seguidores de @{usuario} (límite: {limite_max})...")
    
    for intento_carga in range(max_intentos_carga):
        print(f"\n🔄 Intento de carga {intento_carga + 1}/{max_intentos_carga}")
        
        # Si no es el primer intento, volver al inicio
        if intento_carga > 0:
            print("🔄 Volviendo al inicio antes del nuevo intento de carga...")
            if navegar_a_perfil_desde_inicio(driver, usuario):
                # Intentar ir directamente a followers
                driver.get(f"https://www.instagram.com/{usuario}/followers/")
                sleep(random.uniform(4, 6))
        
        seguidores_unicos = dict()  # Usar dict para almacenar username y fecha
        scroll_count = 0
        sin_nuevos_seguidores = 0
        max_sin_nuevos = 5  # Máximo de scrolls sin nuevos seguidores antes de parar
        
        # Contadores para estadísticas
        total_links_procesados = 0
        links_filtrados = 0
        
        # Esperar a que aparezca el modal de seguidores
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[role='dialog'], div[style*='overflow-y']"))
            )
            print("✅ Modal de seguidores detectado")
        except:
            print("⚠️ Modal no detectado, continuando...")
            if intento_carga < max_intentos_carga - 1:
                continue  # Reintentar desde el inicio
        
        # Proceso principal de carga
        try:
            # Primero, recolectar todos los usernames sin obtener fechas
            print("🔍 Fase 1: Recolectando usernames de seguidores...")
            usernames_temporales = set()
            
            while len(usernames_temporales) < limite_max:
                scroll_count += 1
                count_anterior = len(usernames_temporales)
                
                print(f"🔄 Scroll #{scroll_count} - Usernames encontrados: {len(usernames_temporales)}")
                
                # Selectores mejorados para encontrar seguidores
                selectores_seguidores = [
                    # Para modal de seguidores
                    "div[role='dialog'] a[href^='/']:not([href='/'])",
                    "[role='dialog'] a[href*='/']:not([href='/'])",
                    # Para contenedor scrolleable
                    "div[style*='overflow'] a[href^='/']:not([href='/'])",
                    # Genérico para usernames
                    "a[href^='/']:not([href='/'])",
                ]
                
                # Buscar nuevos seguidores
                elementos_encontrados = []
                for selector in selectores_seguidores:
                    try:
                        elementos = driver.find_elements(By.CSS_SELECTOR, selector)
                        if elementos:
                            elementos_encontrados.extend(elementos)
                    except Exception:
                        continue
                
                # Remover duplicados de elementos encontrados
                elementos_unicos = []
                hrefs_vistos = set()
                for elemento in elementos_encontrados:
                    href = elemento.get_attribute("href")
                    if href and href not in hrefs_vistos:
                        hrefs_vistos.add(href)
                        elementos_unicos.append(elemento)
                
                # Procesar elementos encontrados con filtros mejorados
                for elemento in elementos_unicos:
                    if len(usernames_temporales) >= limite_max:
                        break
                        
                    try:
                        link = elemento.get_attribute("href")
                        total_links_procesados += 1
                        
                        if link:
                            # Limpiar y extraer username
                            username = limpiar_username(link)
                            
                            # Aplicar validación mejorada
                            if username and es_username_valido(username):
                                # Verificar que no lo hayamos agregado ya
                                if username not in usernames_temporales:
                                    usernames_temporales.add(username)
                                    # Debug: mostrar usernames válidos encontrados
                                    if len(usernames_temporales) % 50 == 0:
                                        print(f"✅ Encontrado: @{username} (Total: {len(usernames_temporales)})")
                            else:
                                links_filtrados += 1
                                # Debug opcional: mostrar qué se está filtrando
                                if links_filtrados % 20 == 0:
                                    print(f"🚫 Filtrados hasta ahora: {links_filtrados} enlaces no válidos")
                                
                    except Exception:
                        continue
                
                # Verificar si encontramos nuevos seguidores
                nuevos_encontrados = len(usernames_temporales) - count_anterior
                if nuevos_encontrados == 0:
                    sin_nuevos_seguidores += 1
                    print(f"⚠️ No se encontraron nuevos seguidores válidos en este scroll ({sin_nuevos_seguidores}/{max_sin_nuevos})")
                else:
                    sin_nuevos_seguidores = 0
                    print(f"✅ Encontrados {nuevos_encontrados} nuevos seguidores válidos")
                
                # Parar si no hay nuevos seguidores por varios scrolls seguidos
                if sin_nuevos_seguidores >= max_sin_nuevos:
                    print(f"🛑 No se encontraron nuevos seguidores después de {max_sin_nuevos} intentos. Deteniendo...")
                    break
                
                # Hacer scroll hacia abajo en el modal de seguidores
                try:
                    # Buscar el contenedor scrolleable del modal
                    modal_containers = driver.find_elements(By.CSS_SELECTOR, "div[role='dialog']")
                    if modal_containers:
                        # Scroll dentro del modal
                        driver.execute_script(
                            "arguments[0].scrollTop = arguments[0].scrollHeight;", 
                            modal_containers[0]
                        )
                    else:
                        # Fallback: scroll general
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    
                except Exception as e:
                    print(f"⚠️ Error en scroll: {e}")
                
                # Pausa aleatoria para evitar detección
                sleep(random.uniform(2, 4))
                
                # Mostrar progreso cada 10 scrolls
                if scroll_count % 10 == 0:
                    print(f"📊 Progreso: {len(usernames_temporales)}/{limite_max} seguidores válidos")
                    print(f"🔍 Estadísticas: {total_links_procesados} links procesados, {links_filtrados} filtrados")
            
            # Fase 2: Obtener fechas de unión para cada username encontrado
            print(f"\n📅 Fase 2: Obteniendo fechas de unión para {len(usernames_temporales)} usuarios...")
            
            # Cerrar el modal de seguidores primero
            try:
                driver.execute_script("document.body.click();")
                sleep(2)
            except:
                pass
            
            # Convertir set a lista para procesar
            usernames_lista = list(usernames_temporales)
            
            for i, username in enumerate(usernames_lista, 1):
                print(f"📅 Procesando {i}/{len(usernames_lista)}: @{username}")
                fecha = obtener_fecha_union(driver, username)
                seguidores_unicos[username] = fecha
                
                # Pausa entre cada obtención de fecha para evitar detección
                sleep(random.uniform(1, 2))
                
                # Mostrar progreso cada 10 usuarios
                if i % 10 == 0:
                    print(f"📊 Progreso fechas: {i}/{len(usernames_lista)} completados")
            
            # Si llegamos aquí exitosamente, retornar los resultados
            if seguidores_unicos:
                print(f"✅ Carga exitosa en intento {intento_carga + 1}")
                print(f"📊 ESTADÍSTICAS FINALES:")
                print(f"   - Total links procesados: {total_links_procesados}")
                print(f"   - Links filtrados (no válidos): {links_filtrados}")
                print(f"   - Seguidores válidos encontrados: {len(seguidores_unicos)}")
                print(f"   - Fechas de unión obtenidas: {len([f for f in seguidores_unicos.values() if f != 'Fecha no encontrada'])}")
                print(f"   - Tasa de filtrado: {(links_filtrados/total_links_procesados*100):.1f}%" if total_links_procesados > 0 else "   - Tasa de filtrado: 0%")
                return seguidores_unicos
                
        except Exception as e:
            print(f"❌ Error durante la carga en intento {intento_carga + 1}: {e}")
            if intento_carga < max_intentos_carga - 1:
                print("🔄 Reintentando desde el inicio...")
                sleep(random.uniform(5, 8))
                continue
    
    # Si llegamos aquí, todos los intentos fallaron
    print(f"❌ No se pudo cargar seguidores después de {max_intentos_carga} intentos")
    return dict()

# Inicializar el driver
driver = iniciar_driver()
driver.maximize_window()

insta_url = 'https://www.instagram.com/'
driver.get(insta_url)
sleep(3)

# Inicio de sesión
try:
    username_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'username')))
    username_field.send_keys(USERNAME)

    password_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'password')))
    password_field.send_keys(PASSWORD)

    login_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
    login_button.click()

    print("⚠️ Si Instagram solicita verificación, ingrésala manualmente en el navegador.")
    input("Presiona Enter cuando hayas completado el login/verificación en el navegador...")

    sleep(5)
except Exception as e:
    print(f"Error durante el inicio de sesión: {e}")
    driver.quit()
    exit()

# Verificar que estamos en la página principal después del login
current_url = driver.current_url
if 'instagram.com' not in current_url:
    print("🏠 Navegando a la página principal de Instagram...")
    driver.get('https://www.instagram.com/')
    sleep(3)

# Navegar al perfil del usuario especificado desde el inicio
if not navegar_a_perfil_desde_inicio(driver, usuario_objetivo):
    print(f"❌ No se pudo navegar al perfil de @{usuario_objetivo}")
    driver.quit()
    exit()

# Verificar si el perfil existe
try:
    # Verificar si la página cargó correctamente
    page_title = driver.title.lower()
    current_url = driver.current_url.lower()
    
    # Verificar si es una página de error o perfil no encontrado
    if "page not found" in page_title or "sorry" in page_title or "error" in current_url:
        print(f"❌ Error: El usuario @{usuario_objetivo} no existe o no es público")
        driver.quit()
        exit()
    
except Exception as e:
    print(f"❌ Error al verificar el perfil de @{usuario_objetivo}: {e}")
    print("💡 Verifica que el nombre de usuario esté escrito correctamente")
    driver.quit()
    exit()

# Buscar y hacer clic en el botón de seguidores (con reintentos desde inicio)
print("🔍 Buscando el botón de seguidores...")

seguidores_button = encontrar_boton_seguidores(driver, usuario_objetivo, max_intentos=3)

if seguidores_button:
    try:
        # Scroll hasta el elemento para asegurar que esté visible
        driver.execute_script("arguments[0].scrollIntoView(true);", seguidores_button)
        sleep(2)
        
        # Intentar clic normal primero
        try:
            seguidores_button.click()
            print("✅ Clic exitoso en botón de seguidores")
        except Exception:
            # Si falla, usar JavaScript click
            print("🔄 Intentando clic con JavaScript...")
            driver.execute_script("arguments[0].click();", seguidores_button)
            print("✅ Clic con JavaScript exitoso")
        
        sleep(5)
        
    except Exception as e:
        print(f"❌ Error al hacer clic en seguidores: {e}")
        print("🔄 Navegando directamente a la URL de seguidores...")
        driver.get(f"https://www.instagram.com/{usuario_objetivo}/followers/")
        sleep(5)
else:
    print("❌ No se pudo encontrar el botón de seguidores después de todos los intentos")
    print("🔄 Último intento: navegando directamente a la URL de seguidores...")
    # Volver al inicio antes del último intento
    navegar_a_perfil_desde_inicio(driver, usuario_objetivo)
    sleep(2)
    driver.get(f"https://www.instagram.com/{usuario_objetivo}/followers/")
    sleep(5)
# AGREGAR ESTOS IMPORTS AL INICIO DE TU ARCHIVO (después de los otros imports)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# AGREGAR ESTAS FUNCIONES DESPUÉS DE tus funciones existentes (antes del código principal)
def generar_excel_completo(seguidores_data, usuario_objetivo):
    """
    Genera un archivo Excel completo con los datos de seguidores y estadísticas
    """
    try:
        # Crear DataFrame con los datos
        datos_para_excel = []
        for i, (username, fecha) in enumerate(seguidores_data.items(), 1):
            datos_para_excel.append({
                "No.": i,
                "Usuario": username,
                "Perfil URL": f"https://www.instagram.com/{username}/",
                "Fecha de Unión": fecha,
                "Estado Fecha": "✅ Obtenida" if fecha != "Fecha no encontrada" else "❌ No disponible"
            })
        
        df_seguidores = pd.DataFrame(datos_para_excel)
        
        # Crear workbook y hojas
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # HOJA 1: Datos principales
        ws_datos = wb.create_sheet("Seguidores")
        
        # Añadir título
        ws_datos.merge_cells('A1:E1')
        titulo_cell = ws_datos['A1']
        titulo_cell.value = f"SEGUIDORES DE @{usuario_objetivo.upper()}"
        titulo_cell.font = Font(size=16, bold=True, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Añadir información de extracción
        ws_datos.merge_cells('A2:E2')
        info_cell = ws_datos['A2']
        info_cell.value = f"Extraído el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Total: {len(seguidores_data)} seguidores"
        info_cell.font = Font(size=11, italic=True)
        info_cell.alignment = Alignment(horizontal="center")
        
        # Añadir headers con estilo
        headers = ["No.", "Usuario", "Perfil URL", "Fecha de Unión", "Estado Fecha"]
        for col, header in enumerate(headers, 1):
            cell = ws_datos.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Añadir datos
        for row_idx, row_data in enumerate(df_seguidores.itertuples(index=False), 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_datos.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Estilo alternado para filas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Color especial para URLs
                if col_idx == 3:  # Columna de URLs
                    cell.font = Font(color="0563C1", underline="single")
                
                # Color para estado de fecha
                if col_idx == 5:  # Columna de estado
                    if "✅" in str(value):
                        cell.font = Font(color="70AD47")
                    else:
                        cell.font = Font(color="C5504B")
        
        # Ajustar ancho de columnas
        column_widths = [8, 20, 35, 20, 18]
        for i, width in enumerate(column_widths, 1):
            ws_datos.column_dimensions[chr(64 + i)].width = width
        
        # Añadir bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws_datos.iter_rows(min_row=4, max_row=len(seguidores_data) + 4, 
                                     min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
        
        # HOJA 2: Estadísticas
        ws_stats = wb.create_sheet("Estadísticas")
        
        # Calcular estadísticas
        fechas_exitosas = len([f for f in seguidores_data.values() if f != 'Fecha no encontrada'])
        porcentaje_fechas = (fechas_exitosas / len(seguidores_data) * 100) if seguidores_data else 0
        
        # Añadir estadísticas
        stats_data = [
            ["ESTADÍSTICAS DE EXTRACCIÓN", ""],
            ["", ""],
            ["Usuario analizado:", f"@{usuario_objetivo}"],
            ["Fecha de extracción:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ["Total seguidores extraídos:", len(seguidores_data)],
            ["Fechas de unión obtenidas:", fechas_exitosas],
            ["Fechas no disponibles:", len(seguidores_data) - fechas_exitosas],
            ["Porcentaje de éxito (fechas):", f"{porcentaje_fechas:.1f}%"],
            ["", ""],
            ["FILTROS APLICADOS:", ""],
            ["✅ Enlaces con parámetros eliminados", ""],
            ["✅ IDs temporales eliminados", ""],
            ["✅ Páginas especiales filtradas", ""],
            ["✅ Formato de username validado", ""],
            ["✅ Duplicados eliminados", ""],
        ]
        
        for row_idx, (key, value) in enumerate(stats_data, 1):
            ws_stats.cell(row=row_idx, column=1).value = key
            ws_stats.cell(row=row_idx, column=2).value = value
            
            # Estilo para títulos
            if "ESTADÍSTICAS" in key or "FILTROS" in key:
                ws_stats.cell(row=row_idx, column=1).font = Font(size=14, bold=True, color="FFFFFF")
                ws_stats.cell(row=row_idx, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                ws_stats.merge_cells(f'A{row_idx}:B{row_idx}')
            elif key and "✅" not in key and key != "":
                ws_stats.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Ajustar columnas de estadísticas
        ws_stats.column_dimensions['A'].width = 35
        ws_stats.column_dimensions['B'].width = 25
        
        # HOJA 3: Seguidores con fechas válidas
        seguidores_con_fecha = {k: v for k, v in seguidores_data.items() if v != "Fecha no encontrada"}
        
        if seguidores_con_fecha:
            ws_fechas = wb.create_sheet("Con Fechas de Unión")
            
            # Título
            ws_fechas.merge_cells('A1:D1')
            titulo_fechas = ws_fechas['A1']
            titulo_fechas.value = f"SEGUIDORES CON FECHA DE UNIÓN ({len(seguidores_con_fecha)})"
            titulo_fechas.font = Font(size=14, bold=True, color="FFFFFF")
            titulo_fechas.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            titulo_fechas.alignment = Alignment(horizontal="center")
            
            # Headers
            headers_fechas = ["No.", "Usuario", "Perfil URL", "Fecha de Unión"]
            for col, header in enumerate(headers_fechas, 1):
                cell = ws_fechas.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for i, (username, fecha) in enumerate(seguidores_con_fecha.items(), 1):
                ws_fechas.cell(row=i+3, column=1).value = i
                ws_fechas.cell(row=i+3, column=2).value = username
                ws_fechas.cell(row=i+3, column=3).value = f"https://www.instagram.com/{username}/"
                ws_fechas.cell(row=i+3, column=4).value = fecha
                
                # Estilo alternado
                if (i+3) % 2 == 0:
                    for col in range(1, 5):
                        ws_fechas.cell(row=i+3, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Ajustar columnas
            ws_fechas.column_dimensions['A'].width = 8
            ws_fechas.column_dimensions['B'].width = 20
            ws_fechas.column_dimensions['C'].width = 35
            ws_fechas.column_dimensions['D'].width = 20
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"seguidores_{usuario_objetivo}_{timestamp}.xlsx"
        wb.save(filename)
        
        return filename, True
        
    except Exception as e:
        print(f"❌ Error al generar Excel: {e}")
        return None, False

def mostrar_resumen_final(seguidores_data, usuario_objetivo, filename=None):
    """
    Muestra un resumen final completo de la extracción
    """
    print("\n" + "="*80)
    print("🎉 ¡EXTRACCIÓN COMPLETADA EXITOSAMENTE!")
    print("="*80)
    
    fechas_exitosas = len([f for f in seguidores_data.values() if f != 'Fecha no encontrada'])
    porcentaje_fechas = (fechas_exitosas / len(seguidores_data) * 100) if seguidores_data else 0
    
    print(f"👤 Usuario analizado: @{usuario_objetivo}")
    print(f"📊 Total seguidores extraídos: {len(seguidores_data)}")
    print(f"📅 Fechas de unión obtenidas: {fechas_exitosas} ({porcentaje_fechas:.1f}%)")
    print(f"❌ Fechas no disponibles: {len(seguidores_data) - fechas_exitosas}")
    print(f"⏰ Hora de finalización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    if filename:
        print(f"📄 Archivo Excel generado: {filename}")
        print(f"📁 Ubicación: {os.path.abspath(filename)}")
    
    print("\n🔧 FILTROS APLICADOS:")
    filtros = [
        "✅ Enlaces con parámetros eliminados (?entrypoint, etc.)",
        "✅ IDs temporales filtrados (DJ4EA7ZIaj7, etc.)",
        "✅ Páginas especiales excluidas (followers, inbox, etc.)",
        "✅ Formato de username validado",
        "✅ Duplicados eliminados",
        "✅ Fechas de unión extraídas"
    ]
    
    for filtro in filtros:
        print(f"   {filtro}")
    
    print("="*80)

# REEMPLAZAR TODA LA SECCIÓN desde "# Verificar si estamos en la página de seguidores" CON ESTO:

# Verificar si estamos en la página de seguidores
current_url = driver.current_url
if 'followers' in current_url:
    print("✅ Estamos en la página de seguidores")
    
    # Cargar todos los seguidores (con reintentos desde inicio)
    try:
        seguidores_data = cargar_seguidores_completo(driver, usuario_objetivo, limite_max=100, max_intentos_carga=2)  # Reducido para pruebas
        
        if seguidores_data:
            print(f"\n🎉 ¡Carga completada! Total de seguidores válidos de @{usuario_objetivo}: {len(seguidores_data)}")
            print("\n" + "="*60)
            print(f"📋 LISTA COMPLETA DE SEGUIDORES VÁLIDOS DE @{usuario_objetivo.upper()}:")
            print("="*60)
            
            # Mostrar todos los seguidores encontrados con sus fechas
            for i, (username, fecha) in enumerate(seguidores_data.items(), 1):
                print(f"{i:4d}. 👤 @{username} - {fecha}")
            
            print("="*60)
            print(f"📊 RESUMEN FINAL: {len(seguidores_data)} seguidores válidos extraídos de @{usuario_objetivo}")
            fechas_exitosas = len([f for f in seguidores_data.values() if f != 'Fecha no encontrada'])
            print(f"📅 Fechas de unión obtenidas: {fechas_exitosas}/{len(seguidores_data)} ({fechas_exitosas/len(seguidores_data)*100:.1f}%)")
            print("🔧 FILTROS APLICADOS:")
            print("   ✅ Eliminados enlaces con parámetros (?entrypoint, etc.)")
            print("   ✅ Eliminados IDs temporales (DJ4EA7ZIaj7, etc.)")
            print("   ✅ Eliminadas páginas especiales (followers, inbox, etc.)")
            print("   ✅ Validación de formato de username")
            print("   ✅ Eliminación de duplicados")
            print("   ✅ Obtención de fechas de unión")
            
            # NUEVA FUNCIONALIDAD: Generar archivo Excel
            print("\n📊 Generando archivo Excel...")
            filename, excel_exitoso = generar_excel_completo(seguidores_data, usuario_objetivo)
            
            if excel_exitoso and filename:
                print(f"✅ Archivo Excel generado exitosamente: {filename}")
                
                # Mostrar resumen final
                mostrar_resumen_final(seguidores_data, usuario_objetivo, filename)
                
                # Mostrar preview de los primeros 10 seguidores
                print(f"\n📋 PREVIEW - PRIMEROS 10 SEGUIDORES:")
                print("-" * 60)
                for i, (username, fecha) in enumerate(list(seguidores_data.items())[:10], 1):
                    print(f"{i:2d}. 👤 @{username:<20} | 📅 {fecha}")
                
                if len(seguidores_data) > 10:
                    print(f"    ... y {len(seguidores_data) - 10} más en el archivo Excel")
                
            else:
                print("❌ Error al generar archivo Excel")
                
                # Fallback: guardar en TXT como respaldo
                print("💾 Guardando en archivo de texto como respaldo...")
                try:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    txt_filename = f"seguidores_backup_{usuario_objetivo}_{timestamp}.txt"
                    with open(txt_filename, "w", encoding="utf-8") as f:
                        f.write(f"Seguidores de @{usuario_objetivo} - Total: {len(seguidores_data)}\n")
                        f.write(f"Extraído el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
                        for i, (username, fecha) in enumerate(seguidores_data.items(), 1):
                            f.write(f"{i}. @{username} - {fecha}\n")
                    print(f"✅ Archivo de respaldo guardado: {txt_filename}")
                except Exception as e:
                    print(f"❌ Error al guardar archivo de respaldo: {e}")
        else:
            print("❌ No se pudieron extraer seguidores válidos")
            
    except Exception as e:
        print(f"❌ Error durante la carga de seguidores: {e}")
        
else:
    print("❌ No estamos en la página de seguidores")
    print(f"URL actual: {current_url}")
    print("💡 Intentando navegar directamente a seguidores...")
    driver.get(f"https://www.instagram.com/{usuario_objetivo}/followers/")
    sleep(5)
    # Podrías repetir el proceso aquí si quieres un reintento

print("\n🔚 Proceso finalizado")
print("💡 Tip: Puedes abrir el archivo Excel para ver todos los datos organizados")

# Cerrar driver
driver.quit()
