
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generar_excel_completo(seguidores_data, usuario_objetivo):
    try:
        datos_para_excel = []
        for i, (username, fecha) in enumerate(seguidores_data.items(), 1):
            datos_para_excel.append({
                "No.": i,
                "Usuario": username,
                "Perfil URL": f"https://www.instagram.com/{username}/",
                "Fecha de Unión": fecha,
                "Estado Fecha": "✅ Obtenida" if fecha != "Fecha no encontrada" else "❌ No disponible"
            })

        df = pd.DataFrame(datos_para_excel)
        wb = Workbook()
        wb.remove(wb.active)

        # Hoja 1: Seguidores
        ws = wb.create_sheet("Seguidores")
        ws.merge_cells('A1:E1')
        ws['A1'] = f"SEGUIDORES DE @{usuario_objetivo.upper()}"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill("solid", fgColor="366092")
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:E2')
        ws['A2'] = f"Extraído el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Total: {len(seguidores_data)} seguidores"
        ws['A2'].alignment = Alignment(horizontal="center")
        ws['A2'].font = Font(italic=True)

        headers = ["No.", "Usuario", "Perfil URL", "Fecha de Unión", "Estado Fecha"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(df.itertuples(index=False), 5):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                if col_idx == 3:
                    cell.font = Font(color="0563C1", underline="single")
                if col_idx == 5:
                    cell.font = Font(color="70AD47" if "✅" in str(value) else "C5504B")

        for i, w in enumerate([8, 20, 35, 20, 18], 1):
            ws.column_dimensions[chr(64+i)].width = w

        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=4, max_row=len(df) + 4, min_col=1, max_col=5):
            for cell in row:
                cell.border = border

        # Hoja 2: Estadísticas
        ws2 = wb.create_sheet("Estadísticas")
        fechas_ok = sum(1 for f in seguidores_data.values() if f != "Fecha no encontrada")
        stats = [
            ["Usuario analizado:", f"@{usuario_objetivo}"],
            ["Fecha de extracción:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ["Total seguidores extraídos:", len(seguidores_data)],
            ["Fechas de unión obtenidas:", fechas_ok],
            ["Fechas no disponibles:", len(seguidores_data) - fechas_ok],
            ["Porcentaje de éxito (fechas):", f"{(fechas_ok / len(seguidores_data) * 100):.1f}%"] if seguidores_data else ["", ""]
        ]
        for i, (k, v) in enumerate(stats, 1):
            ws2.cell(row=i, column=1).value = k
            ws2.cell(row=i, column=2).value = v
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 25

        # Hoja 3: Con fechas válidas
        con_fecha = {k: v for k, v in seguidores_data.items() if v != "Fecha no encontrada"}
        if con_fecha:
            ws3 = wb.create_sheet("Con Fechas de Unión")
            headers_fechas = ["No.", "Usuario", "Perfil URL", "Fecha de Unión"]
            for col, header in enumerate(headers_fechas, 1):
                c = ws3.cell(row=1, column=col)
                c.value = header
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="4472C4")

            for i, (u, f) in enumerate(con_fecha.items(), 1):
                ws3.cell(row=i+1, column=1).value = i
                ws3.cell(row=i+1, column=2).value = u
                ws3.cell(row=i+1, column=3).value = f"https://www.instagram.com/{u}/"
                ws3.cell(row=i+1, column=4).value = f

            for i, w in zip('ABCD', [8, 20, 35, 20]):
                ws3.column_dimensions[i].width = w

        filename = f"seguidores_{usuario_objetivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename, True

    except Exception as e:
        print(f"❌ Error al generar Excel: {e}")
        return None, False


def mostrar_resumen_final(seguidores_data, usuario_objetivo, filename=None):
    print("\n" + "="*80)
    print("🎉 ¡EXTRACCIÓN COMPLETADA EXITOSAMENTE!")
    print("="*80)
    fechas_ok = sum(1 for f in seguidores_data.values() if f != 'Fecha no encontrada')
    print(f"👤 Usuario analizado: @{usuario_objetivo}")
    print(f"📊 Total seguidores extraídos: {len(seguidores_data)}")
    print(f"📅 Fechas de unión obtenidas: {fechas_ok} ({(fechas_ok / len(seguidores_data) * 100):.1f}%)")
    print(f"❌ Fechas no disponibles: {len(seguidores_data) - fechas_ok}")
    print(f"⏰ Hora de finalización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    if filename:
        print(f"📄 Archivo Excel generado: {filename}")
        print(f"📁 Ubicación: {os.path.abspath(filename)}")
    print("\n🔧 FILTROS APLICADOS:")
    for filtro in [
        "✅ Enlaces con parámetros eliminados",
        "✅ IDs temporales eliminados",
        "✅ Páginas especiales filtradas",
        "✅ Formato de username validado",
        "✅ Duplicados eliminados",
        "✅ Fechas de unión extraídas"
    ]:
        print(f"   {filtro}")
    print("="*80)
